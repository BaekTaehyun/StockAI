import openpyxl

wb = openpyxl.load_workbook('키움 REST API 문서.xlsx')

# ka20019 시트 찾기
sheet_name = '업종년봉조회(ka20019)'
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Sheet: {sheet_name}")
    print("="*60)
    
    # 전체 시트 출력
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 60:
            break
        if any(cell for cell in row):  # 빈 행 건너뛰기
            print(f"{i}: {row}")
