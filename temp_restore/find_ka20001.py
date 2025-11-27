import openpyxl

wb = openpyxl.load_workbook('키움 REST API 문서.xlsx')

print("Searching for ka20001 (업종기본정보조회)...")
print("="*60)

for sheet_name in wb.sheetnames:
    if 'ka20001' in sheet_name:
        print(f"\nFound sheet: {sheet_name}\n")
        ws = wb[sheet_name]
        
        # Print first 40 rows
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i > 40:
                break
            if any(cell for cell in row):
                print(f"{i}: {row}")
