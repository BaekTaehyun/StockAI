import openpyxl

wb = openpyxl.load_workbook('키움 REST API 문서.xlsx')

print("Searching for ka20019 details...")
print("="*60)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_str = ' '.join([str(cell) if cell else '' for cell in row])
        if 'ka20019' in row_str.lower():
            print(f"\n[{sheet_name}] Row {i}:")
            for j, cell in enumerate(row):
                if cell:
                    print(f"  Col {j}: {cell}")
