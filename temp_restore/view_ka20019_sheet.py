import openpyxl

wb = openpyxl.load_workbook('키움 REST API 문서.xlsx')

# ka20019 시트 찾기
if '업종년봉조회(ka20019)' in wb.sheetnames:
    ws = wb['업종년봉조회(ka20019)']
    print("Sheet: 업종년봉조회(ka20019)")
    print("="*60)
    
    # 전체 시트 출력 (처음 50행)
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 50:
            break
        if any(cell for cell in row):  # 빈 행 건너뛰기
            print(f"Row {i}: {row}")
