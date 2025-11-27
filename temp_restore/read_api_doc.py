import openpyxl

# Excel 파일 열기
wb = openpyxl.load_workbook('키움 REST API 문서.xlsx')

print("=== Sheet Names ===")
print(wb.sheetnames)
print()

# 각 시트 확인
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print('='*60)
    
    # 첫 30행 출력
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 30:
            break
        print(f"Row {i}: {row}")
    print()

# "지수" 또는 "업종" 키워드 검색
print("\n" + "="*60)
print("Searching for '지수' or '업종' or 'index'...")
print("="*60)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_str = ' '.join([str(cell) if cell else '' for cell in row])
        if '지수' in row_str or '업종' in row_str or 'index' in row_str.lower():
            print(f"\n[{sheet_name}] Row {i}: {row}")
